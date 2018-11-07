#!/usr/bin/env python

from __future__ import print_function, division
import json
import sys
import os
import argparse
from collections import defaultdict
import matplotlib.pyplot as plt
import lxml.etree as ET
import networkx as nx
from openpyxl import Workbook
from fnmatch import fnmatch


def first(x):
    if x:
        return x[0]
    return None


def load_json(filename):
    with open(filename) as fp:
        return json.load(fp)
        
    
def read_graphml(stream):
    """
    Read factory graph from graphml
    """
    tree = ET.parse(stream)
        
    ns = {
        'g': "http://graphml.graphdrawing.org/xmlns",
        'y': "http://www.yworks.com/xml/graphml",
    }
    graph = first(tree.xpath("/g:graphml/g:graph", namespaces=ns))
    assert graph is not None
    
    nodes_data = []
    nodes = graph.xpath("g:node", namespaces=ns)
    for x in nodes:
        label = first(x.xpath(".//y:NodeLabel/text()", namespaces=ns))
        nodes_data.append((x.get('id'), label))
        
    edges_data = []
    edges = graph.xpath("g:edge", namespaces=ns)
    for x in edges:
        label = first(x.xpath(".//y:EdgeLabel/text()", namespaces=ns))
        edges_data.append((x.get('id'), x.get('source'), x.get('target'), label))
        
    return nodes_data, edges_data
    
    
def detect_cycles(G, start):
    visited = set()
    visited.add(start)

    stack = [(start, iter(G[start]))]
    while stack:
        parent, children = stack[-1]
        try:
            child = next(children)
            if child not in visited:
                yield parent, child
                visited.add(child)
                stack.append((child, iter(G[child])))
            else:
                print("Visited:", child)
        except StopIteration:
            stack.pop()
            
            
def read_production_graph(filename):
    with open(filename) as fp:
        nodes, edges = read_graphml(fp)
        
    node_cnt = defaultdict(int)
    d_nodes = {}
    for n in nodes:
        d_nodes[n[0]] = n[1], node_cnt[n[1]]
        node_cnt[n[1]] += 1
        
    g = nx.DiGraph()
    for n in nodes:
        node_recipe_name = d_nodes[n[0]]
        g.add_node(d_nodes[n[0]], {'label': node_recipe_name})
        
    for e in edges:
        g.add_edge(d_nodes[e[1]], d_nodes[e[2]], {'label': e[3]})
        
    return g


DEFAULT_CRAFTING_SPEEDS = {
    'chemistry': 1.25,
    'liquifying': 2.25, #1.5,
    'crafting': 0.75,  # 0.5
    'advanced-crafting': 0.75,
    'ore-sorting-t1': 1.5,
    'petrochem-electrolyser': 1.5, #1,
    'smelting': 1,
    'bio-processing': 1.5, #1,
    'filtering': 1.5,
    'washing-plant': 1.5,
    'electronics': 2.25,
    'electronics-machine': 2.25,
}


class Recipe(object):
    @staticmethod
    def _unpack_data(lst):
        return {
            x['name']: {
                'name': x['name'],
                'amount': x['amount'],
                'per_second': None,
                'per_second_real': None
            }
            for x in lst
        }

    def __init__(self, data):
        self.name = data['name']
        self.time = data['energy']
        self.category = data['category']
        self.ingredients = Recipe._unpack_data(data['ingredients'])
        self.products = Recipe._unpack_data(data['products'])
        self.num_machines = 1.0
        self.crafting_speed = DEFAULT_CRAFTING_SPEEDS.get(self.category, 1.0)
        
    def update(self):
        per_second_1 = self.crafting_speed / self.time
        per_second_1_real = per_second_1 * self.num_machines
        for collection in (self.ingredients, self.products):
            for x in collection.itervalues():
                x['per_second'] = x['amount'] * per_second_1
                x['per_second_real'] = x['amount'] * per_second_1_real
                
                
def dump_excel(filename, graph, nodes, edges):
    wb = Workbook()
    ws = wb.active
    ws.title = 'recipes'
    ws.append(["name", "amount", "time", "per-second-raw", "crafting-speed", "num-machines", "per-second-real"])
    ws.append([])
    
    totals = defaultdict(list)
    heads = {}
    prod_rows = {}

    for n in nodes:
        r = graph.node[n]['r']
        ws.append([r.name, r.category, r.time, None, r.crafting_speed, r.num_machines])
        head = ws.max_row
        heads[n] = head
        for ip in ('products', 'ingredients'):
            for i in getattr(r, ip).values():
                amount = i['amount']
                this_row = ws.max_row + 1
                mul = -1 if ip == 'ingredients' else 1
                i_name = i['name']
                prod_rows[(n, i_name)] = this_row
                x = ws.append({
                    'A': i['name'],
                    'B': mul * i['amount'],
                    'D': '=B{1}/C{0}*E{0}'.format(head, this_row),
                    'G': '=F{0}*D{1}'.format(head, this_row),
                })
                totals[i['name']].append('G{0}'.format(this_row))
        ws.append([])
        
    for e_src, e_dst in edges:
        # print(e_src, e_dst)
        common = common_io(graph, e_src, e_dst)
        p_src = prod_rows[(e_src, common)]
        p_dst = prod_rows[(e_dst, common)]
        ws['F{0}'.format(heads[e_dst])] = '=-G{0}/D{1}'.format(p_src, p_dst)
    
    ws.append([])
    for name, used_in_cells in totals.iteritems():
        ws.append({
            'A': name,
            'G': '=' + '+'.join(used_in_cells),
        })
    
    wb.save(filename)
    
    
def common_io(graph, a, b):
    a_out = frozenset(graph.node[a]['r'].ingredients.keys())
    b_in = frozenset(graph.node[b]['r'].products.keys())
    common = a_out & b_in
    if len(common) == 0:
        raise RuntimeError("Recipes mismatch {} {}".format(a, b))
    if len(common) >= 2:
        raise NotImplementedError("Two common ingredients is not supported")
    return list(common)[0]
    
    
def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input-dir', '-i', default='.')
    return parser.parse_args()

                
def main():
    # TODO:
    # - multiple outputs
    # - cycles

    args = parse_args()
    recipes = load_json('recipes.json')
    input_dir = args.input_dir
    input_file = [file for file in os.listdir(input_dir) if fnmatch(file, '*.graphml')][0]
    input_file = os.path.join(input_dir, input_file)
    output_file = os.path.splitext(input_file)[0] + '.xlsx'
    
    graph = read_production_graph(input_file)
    target_node = [x for x in graph.nodes() if graph.out_degree(x) == 0][0]
    
    rev_graph = graph.reverse(copy=True)
    
    x_nodes = list(nx.dfs_preorder_nodes(rev_graph, target_node))
    x_edges = list(nx.bfs_edges(graph, target_node, reverse=True))

    for r_name in graph.nodes():
        r = Recipe(recipes[r_name[0]])
        r.update()
        graph.node[r_name]['r'] = r

    dump_excel(output_file, graph, x_nodes, x_edges)

    
if __name__ == '__main__':
    main()